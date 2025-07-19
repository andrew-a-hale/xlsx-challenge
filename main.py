#!/usr/bin/env -S uv run --script
# /// script
# dependencies = ["duckdb", "openpyxl", "typer"]
# ///
import enum
import json
import os
import random
import tempfile
import zipfile
from typing import Optional

import duckdb
import typer
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Dep(enum.Enum):
    ENGINEERING = 1
    SALES = 2
    MARKETING = 3
    FINANCE = 4
    HUMAN_RESOURCES = 5


class EmpType(enum.Enum):
    FULL_TIME = 1
    PART_TIME = 2
    CASUAL = 3
    CONTRACT = 4
    TEMPORARY = 5


class Unit(enum.Enum):
    BRISBANE = 1
    SYDNEY = 2
    MELBOURNE = 3


class Position(enum.Enum):
    SOFTWARE_ENGINEER = 1
    SALES_REPRESENTATIVE = 2
    MARKETING_SPECIALIST = 3
    FINANCIAL_ANALYST = 4
    HR_COORDINATOR = 5
    PRODUCT_MANAGER = 6
    DATA_SCIENTIST = 7
    CUSTOMER_SUPPORT = 8
    OPERATIONS_MANAGER = 9
    SENIOR_DEVELOPER = 10


UNIT_SCHEMAS = {
    "BRISBANE": [
        "row_id",
        "first_name",
        "last_name",
        "dob",
        "emp_start",
        "date",
        "position",
        "emp_type",
        "hourly_rate",
    ],
    "SYDNEY": [
        "row_id",
        "full_name_one",
        "dob",
        "date",
        "position",
        "hourly_rate",
        "emp_start",
        "emp_type",
    ],
    "MELBOURNE": [
        "row_id",
        "full_name_two",
        "dob",
        "emp_start",
        "date",
        "emp_type",
        "hourly_rate",
        "position",
    ],
}


def unit_fn(unit: str, fortnight: int) -> str:
    match unit:
        case Unit.BRISBANE.name:
            return f"fn {fortnight}"
        case Unit.SYDNEY.name:
            return f"fortnight-{fortnight}"
        case Unit.MELBOURNE.name:
            return f"period_{fortnight}"

    x = random.randint(0, 10)
    return f"overflow_{x}"


def unit_wb(unit: str, year: int) -> str:
    match unit:
        case Unit.BRISBANE.name:
            return f"PAYROLL-{year}"
        case Unit.SYDNEY.name:
            return f"PAY {year}"
        case Unit.MELBOURNE.name:
            return f"payroll_{year}"

    x = random.randint(0, 10)
    return f"{unit}_new_unit_{x}"


def add_junk(ws: Worksheet, row_offset: int) -> int:
    rand_rows = random.randint(2, 10)
    rand_cols = random.randint(2, 10)
    for r in range(rand_rows):
        tmp = r + row_offset
        for c in range(rand_cols):
            if c == 0:
                ws.cell(tmp, c + 1, tmp - (row_offset > 1))
            else:
                ws.cell(tmp, c + 1, random.randbytes(random.randint(1, 10)).hex())

    return rand_rows + row_offset


def write_table(
    unit: str,
    ws: Worksheet,
    df: duckdb.DuckDBPyRelation,
    row_offset: int,
) -> int:
    data = df.fetchall()
    for i, header in enumerate(UNIT_SCHEMAS[unit]):
        ws.cell(row_offset, i + 1, header)

    row_offset += 1
    for r, _ in enumerate(data):
        tmp = r + row_offset
        for c, col in enumerate(UNIT_SCHEMAS[unit]):
            if c == 0:
                ws.cell(tmp, c + 1, tmp - 1)
            else:
                col_index = df.columns.index(col)
                ws.cell(tmp, c + 1, data[r][col_index])

    return row_offset + len(df)


def write_sheet(
    unit: str,
    year: int,
    fortnight: int,
    wb: Workbook,
    df: duckdb.DuckDBPyRelation,
):
    print(unit, year, fortnight + 1)
    ws = wb.create_sheet(f"{unit_fn(unit, fortnight)}")
    offset = 1

    # add header junk 50%
    if random.random() < 0.5:
        offset = add_junk(ws, offset)

    # write table
    tmp = df.filter(f"unit == '{unit}' and year == {year} and fortnight == {fortnight}")
    offset = write_table(unit, ws, tmp, offset)

    # duplicate table 2%
    if random.random() < 0.02:
        offset = write_table(unit, ws, tmp, offset)

    # add footer junk 50%
    if random.random() < 0.5:
        add_junk(ws, offset)


def create_data(source: str, out: str):
    outfile = open(tempfile.mktemp(suffix=".ndjson"), mode="w")
    with open(source) as f:
        for line in f:
            j = json.loads(line)
            rand_dep = random.randint(1, 5)
            rand_emp_type = random.randint(1, 5)
            rand_unit = random.randint(1, 3)
            rand_pos = random.randint(1, 10)
            j["department"] = Dep(rand_dep).name
            j["unit"] = Unit(rand_unit).name
            j["emp_type"] = EmpType(rand_emp_type).name
            j["position"] = Position(rand_pos).name
            outfile.write(json.dumps(j))
            outfile.write("\n")
        outfile.flush()
    outfile.close()

    df = duckdb.sql(f"""\
with transform as (
    select 
        first_name,
        last_name,
        first_name || ' ' || last_name as full_name_one,
        last_name || ', ' || first_name as full_name_two,
        dob,
        wage / 52 / 5 / 8 hourly_rate,
        emp_start,
        least(days_employed::int + emp_start::date, today()) as emp_end,
        department,
        unit,
        emp_type,
        position
    from '{outfile.name}'
),

ranges as (
    select min(emp_start) as start_date, today() as end_date
    from transform
),

dates as (
    select s.generate_series as date, dayofyear(date) // 14 as fortnight, year(date) as year
    from ranges
    cross join generate_series(start_date::date, end_date::date, interval '1' day) as s
),

expand as (
    select *
    from transform
    inner join dates on date between emp_start and emp_end
)

select 0 as row_id, *
from expand
order by unit, year, fortnight, department""")

    units = sorted(x[0] for x in df.select("unit").distinct().fetchall())
    years = sorted(x[0] for x in df.select("year").distinct().fetchall())
    fortnights = sorted(x[0] for x in df.select("fortnight").distinct().fetchall())

    duplicates = []
    for unit in units:
        for year in years:
            wb = Workbook()
            for fortnight in fortnights:
                write_sheet(unit, year, fortnight, wb, df)

                # duplicate sheet 5%
                if random.random() < 0.05:
                    duplicates.append((unit, year, fortnight))

            wb.remove(wb["Sheet"])
            wb.save(os.path.join(out, f"{unit_wb(unit, year)}.xlsx"))

    for unit, year, fortnight in duplicates:
        wb = Workbook()
        write_sheet(unit, year, fortnight, wb, df)
        wb.remove(wb["Sheet"])
        wb.save(os.path.join(out, f"{unit_wb(unit, year)}_extra.xlsx"))


def create_zip(out: str):
    zip_filename = f"{out}.zip"

    with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(out):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, start=os.path.dirname(out))
                zipf.write(file_path, arcname=arcname)


def main(
    source: str = typer.Option(
        "payroll_data.ndjson",
        help="Source file or directory",
    ),
    out: Optional[str] = typer.Option(None, help="Output directory for XLSX files"),
):
    if out is None:
        out = "xlsx"
    os.makedirs(out, exist_ok=True)

    create_data(source, out)
    create_zip(out)


if __name__ == "__main__":
    typer.run(main)
